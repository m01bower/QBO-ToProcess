"""
AutoProcess Test - Read AutoProcess config and fetch QBO reports to Excel.

Reads the AutoProcess tab from the Financials Google Sheet, fetches each
configured report from QBO with the specified parameters, applies any
post-processing (comparison interleaving, product filtering), and exports
to Excel files for verification.

Usage:
    python src/test_autoprocess.py --client BostonHCP
    python src/test_autoprocess.py --client BostonHCP --row 3
    python src/test_autoprocess.py --client BostonHCP --dry-run
"""

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import List, Tuple, Any, Dict

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import QBO_REPORTS, MAX_ALL
from settings import load_settings
from services.qbo_service import QBOService
from services.sheets_service import SheetsService
from processors.comparison_processor import (
    interleave_comparison_columns,
    filter_rows_by_products,
)
from logger_setup import setup_logger, get_logger


def get_project_root() -> Path:
    """Get project root from this script's location (src/)."""
    return Path(__file__).parent.parent


def get_export_dir(output_dir: str = "exports") -> Path:
    """Get or create the exports directory."""
    export_path = get_project_root() / output_dir
    export_path.mkdir(parents=True, exist_ok=True)
    return export_path


def sanitize_filename(name: str) -> str:
    """Remove or replace characters invalid in filenames."""
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name


def generate_filename(client_name: str, report_name: str, row_index: int) -> str:
    """Generate filename like: BostonHCP_Row3_P&L_2026-02-13.xlsx"""
    safe_report = sanitize_filename(report_name.replace(" ", ""))
    date_str = date.today().strftime("%Y-%m-%d")
    return f"{client_name}_Row{row_index}_{safe_report}_{date_str}.xlsx"


def write_to_excel(
    filepath: Path,
    headers: List[str],
    rows: List[List[Any]],
    report_name: str,
) -> int:
    """
    Write headers and rows to an Excel file.

    Returns the number of data rows written.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_name[:31]  # Excel tab names max 31 chars

    bold_font = Font(bold=True)

    # Write headers in row 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Write data rows starting at row 2
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell_value = value
            if isinstance(value, str):
                cleaned = value.replace(",", "").strip()
                try:
                    cell_value = float(cleaned)
                    if cell_value == int(cell_value):
                        cell_value = int(cell_value)
                except (ValueError, OverflowError):
                    cell_value = value
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filepath)
    return len(rows)


def print_config(year: int, configs: List[Dict[str, Any]]) -> None:
    """Print the AutoProcess configuration in a readable format."""
    print(f"\n  Year: {year}")
    print(f"  Rows: {len(configs)}")
    print()

    for cfg in configs:
        report = cfg["report_name"]
        validated = report in QBO_REPORTS
        status = "OK" if validated else "UNKNOWN"

        print(f"  Row {cfg['row_index']:>2d}: {report:40s} [{status}]")
        print(f"          date_range={cfg['date_range']:<8s} "
              f"display={cfg['display']:<12s} "
              f"basis={cfg['basis']:<8s}")
        if cfg["special_options"]:
            print(f"          special_options={cfg['special_options']}")
    print()


def run_autoprocess(
    client_name: str,
    row_filter: int = None,
    dry_run: bool = False,
    output_dir: str = "exports",
) -> int:
    """
    Main AutoProcess workflow.

    Returns 0 on success, 1 on any failure.
    """
    logger = get_logger()

    print(f"\n{'=' * 60}")
    print(f"  AutoProcess Test")
    print(f"{'=' * 60}")

    # Load settings
    settings = load_settings()
    if not settings.is_configured():
        print("ERROR: App not configured. Run: python src/main.py --setup")
        return 1

    # Validate client
    if client_name not in settings.clients:
        print(f"ERROR: Unknown client '{client_name}'")
        print(f"Available: {', '.join(settings.clients.keys())}")
        return 1

    client_config = settings.clients[client_name]
    sheet_id = client_config.toprocess_sheet_id
    if not sheet_id:
        print(f"ERROR: No Google Sheet ID configured for {client_name}")
        return 1

    print(f"\n  Client: {client_name}")
    print(f"  Sheet:  {sheet_id}")

    # Authenticate with Google Sheets
    print("\nConnecting to Google Sheets...")
    sheets = SheetsService(
        auth_method=client_config.google_auth_method,
        client_name=client_name,
    )
    if not sheets.authenticate():
        print("ERROR: Google Sheets authentication failed")
        return 1

    # Read AutoProcess config
    print("Reading AutoProcess tab...")
    year, configs = sheets.read_autoprocess_config(sheet_id)

    if year is None:
        print("ERROR: Failed to read AutoProcess configuration")
        return 1

    if not configs:
        print("WARNING: No configurations found in AutoProcess tab")
        return 0

    # Filter to specific row if requested
    if row_filter is not None:
        configs = [c for c in configs if c["row_index"] == row_filter]
        if not configs:
            print(f"ERROR: No configuration found for row {row_filter}")
            return 1

    # Print config
    print_config(year, configs)

    # Dry run - stop here
    if dry_run:
        print("  (dry run - no reports fetched)")
        return 0

    # Connect to QBO
    print("Connecting to QuickBooks...")
    qbo = QBOService(settings.qbo_app, client_name)

    if not qbo.is_authenticated:
        print(f"ERROR: {client_name} not authorized for QBO.")
        print(f"Run: python src/main.py --auth {client_name}")
        return 1

    if not qbo.test_connection():
        print(f"ERROR: Failed to connect to QBO for {client_name}")
        return 1

    print("Connected!\n")

    # Create export directory
    export_dir = get_export_dir(output_dir)
    print(f"Export directory: {export_dir}\n")

    # Process each config row
    results: List[Tuple[int, str, str, int, str]] = []

    for cfg in configs:
        row_idx = cfg["row_index"]
        report_name = cfg["report_name"]
        date_range = cfg["date_range"]
        display = cfg["display"]
        basis = cfg["basis"]
        special_options = cfg["special_options"]

        print(f"Row {row_idx}: {report_name} ({display}, {basis})...")

        # Validate report name
        if report_name not in QBO_REPORTS:
            print(f"  SKIP: Unknown report type '{report_name}'")
            results.append((row_idx, report_name, "SKIP", 0, ""))
            continue

        # Fetch report
        try:
            report_data = qbo.get_report(
                report_name=report_name,
                year=year,
                display=display,
                basis=basis,
                full_year=True,
                date_range=date_range,
                special_options=special_options,
            )
        except Exception as e:
            print(f"  ERROR: {e}")
            logger.exception(f"Failed to fetch {report_name} (row {row_idx})")
            results.append((row_idx, report_name, "FAILED", 0, ""))
            continue

        if not report_data:
            print(f"  ERROR: No data returned")
            results.append((row_idx, report_name, "FAILED", 0, ""))
            continue

        # Parse report to rows
        rows, headers = qbo.parse_report_to_rows(
            report_data,
            row_max=MAX_ALL,
            col_max=MAX_ALL,
        )

        if not rows and not headers:
            print(f"  WARNING: Empty report")
            results.append((row_idx, report_name, "EMPTY", 0, ""))
            continue

        # Apply comparison interleaving if requested
        if "Comparison" in special_options:
            print(f"  Applying comparison interleaving...")
            headers, rows = interleave_comparison_columns(headers, rows, year)

        # Apply product filter if special_options has product names
        # Products are listed as comma-separated names that aren't known flags
        known_flags = {"Comparison"}
        if special_options:
            option_parts = [p.strip() for p in special_options.split(",")]
            product_names = [p for p in option_parts if p and p not in known_flags]
            if product_names:
                print(f"  Filtering to products: {product_names}")
                headers, rows = filter_rows_by_products(headers, rows, product_names)

        # Write to Excel
        filename = generate_filename(client_name, report_name, row_idx)
        filepath = export_dir / filename

        num_rows = write_to_excel(filepath, headers, rows, report_name)
        print(f"  Saved: {filename} ({num_rows} rows, {len(headers)} columns)")
        results.append((row_idx, report_name, "OK", num_rows, filename))

    # Print summary
    error_count = sum(1 for _, _, status, _, _ in results if status in ("FAILED", "SKIP"))

    print(f"\n{'=' * 60}")
    print("  SUMMARY")
    print(f"{'=' * 60}")
    for row_idx, rname, status, nrows, fname in results:
        if status == "OK":
            print(f"  OK    Row {row_idx:>2d}: {rname:35s} {nrows:>5d} rows -> {fname}")
        elif status == "EMPTY":
            print(f"  EMPTY Row {row_idx:>2d}: {rname:35s}")
        elif status == "SKIP":
            print(f"  SKIP  Row {row_idx:>2d}: {rname:35s}")
        else:
            print(f"  FAIL  Row {row_idx:>2d}: {rname:35s}")

    total = len(results)
    ok_count = sum(1 for _, _, s, _, _ in results if s == "OK")
    print(f"\n  {ok_count}/{total} reports exported successfully.")
    if error_count:
        print(f"  {error_count} report(s) failed or skipped.")
    print()

    return 0 if error_count == 0 else 1


def main():
    """Entry point."""
    parser = argparse.ArgumentParser(
        description="AutoProcess Test - Fetch QBO reports based on AutoProcess tab config",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python src/test_autoprocess.py --client BostonHCP
  python src/test_autoprocess.py --client BostonHCP --row 3
  python src/test_autoprocess.py --client BostonHCP --dry-run
        """,
    )

    parser.add_argument("--client", required=True, help="Client name")
    parser.add_argument(
        "--row", type=int, default=None,
        help="Process only a specific row number from AutoProcess tab",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Read config only, don't fetch reports",
    )
    parser.add_argument(
        "--output-dir", default="exports",
        help="Output directory relative to project root (default: exports)",
    )

    args = parser.parse_args()

    # Console-only logging for test script
    setup_logger(log_to_file=False)

    exit_code = run_autoprocess(
        client_name=args.client,
        row_filter=args.row,
        dry_run=args.dry_run,
        output_dir=args.output_dir,
    )
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
