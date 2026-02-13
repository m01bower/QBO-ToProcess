"""
QBO Test Export - Download QBO reports to Excel files.

Diagnostic script that connects to QuickBooks Online using existing
cached OAuth tokens, downloads reports, and saves them as Excel files.
No Google Sheets interaction.

Usage:
    python src/test_qbo_export.py --client BostonHCP
    python src/test_qbo_export.py --client BostonHCP --report "Balance Sheet" --report "P&L"
    python src/test_qbo_export.py --client BostonHCP --year 2025 --basis Cash
    python src/test_qbo_export.py --list-clients
    python src/test_qbo_export.py --list-reports
"""

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import List, Tuple, Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import QBO_REPORTS, MAX_ALL
from settings import load_settings
from services.qbo_service import QBOService
from logger_setup import setup_logger, get_logger

# Default reports to fetch when none specified
DEFAULT_REPORTS = [
    "Balance Sheet",
    "P&L",
    "AR Aging",
    "Sales by Customer Summary",
]


def get_project_root() -> Path:
    """Get project root from this script's location (src/)."""
    return Path(__file__).parent.parent


def get_export_dir(output_dir: str = "exports") -> Path:
    """Get or create the exports directory."""
    export_path = get_project_root() / output_dir
    export_path.mkdir(parents=True, exist_ok=True)
    return export_path


def sanitize_filename(name: str) -> str:
    """Remove or replace characters invalid in filenames."""
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name


def generate_filename(client_name: str, report_name: str) -> str:
    """Generate filename like: BostonHCP_BalanceSheet_2026-02-12.xlsx"""
    safe_report = sanitize_filename(report_name.replace(" ", ""))
    date_str = date.today().strftime("%Y-%m-%d")
    return f"{client_name}_{safe_report}_{date_str}.xlsx"


def write_to_excel(
    filepath: Path,
    headers: List[str],
    rows: List[List[Any]],
    report_name: str,
) -> int:
    """
    Write headers and rows to an Excel file.

    Returns the number of data rows written.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_name[:31]  # Excel tab names max 31 chars

    bold_font = Font(bold=True)

    # Write headers in row 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Write data rows starting at row 2
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell_value = value
            if isinstance(value, str):
                cleaned = value.replace(",", "").strip()
                try:
                    cell_value = float(cleaned)
                    if cell_value == int(cell_value):
                        cell_value = int(cell_value)
                except (ValueError, OverflowError):
                    cell_value = value
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filepath)
    return len(rows)


def list_clients():
    """Print configured clients and their status."""
    settings = load_settings()
    print("\nConfigured clients:")
    print("-" * 50)
    for name, cfg in settings.clients.items():
        status = "enabled" if cfg.enabled else "disabled"
        realm = cfg.qbo_realm_id or "(no realm ID)"
        print(f"  {name:20s} [{status:>8s}]  realm={realm}")
    if not settings.clients:
        print("  (none configured -- run main.py --setup first)")
    print()


def list_reports():
    """Print available QBO report types."""
    print("\nAvailable QBO reports:")
    print("-" * 50)
    for display_name, api_name in QBO_REPORTS.items():
        default_marker = " *" if display_name in DEFAULT_REPORTS else ""
        print(f"  {display_name:30s} -> {api_name}{default_marker}")
    print("\n  * = included in default set\n")


def export_reports(
    client_name: str,
    report_names: List[str],
    year: int,
    display: str,
    basis: str,
    output_dir: str,
) -> int:
    """
    Main export workflow.

    Returns 0 on success, 1 on any failure.
    """
    logger = get_logger()

    print(f"\n{'=' * 60}")
    print(f"  QBO Test Export")
    print(f"{'=' * 60}")

    # Load settings
    settings = load_settings()
    if not settings.is_configured():
        print("ERROR: App not configured. Run: python src/main.py --setup")
        return 1

    # Validate client
    if client_name not in settings.clients:
        print(f"ERROR: Unknown client '{client_name}'")
        print(f"Available: {', '.join(settings.clients.keys())}")
        return 1

    client_config = settings.clients[client_name]
    if not client_config.enabled:
        print(f"WARNING: Client '{client_name}' is disabled, proceeding anyway...")

    # Validate report names
    for rname in report_names:
        if rname not in QBO_REPORTS:
            print(f"ERROR: Unknown report '{rname}'")
            print(f"Available: {', '.join(QBO_REPORTS.keys())}")
            return 1

    print(f"\n  Client:  {client_name}")
    print(f"  Year:    {year}")
    print(f"  Display: {display}")
    print(f"  Basis:   {basis}")
    print(f"  Reports: {', '.join(report_names)}")
    print()

    # Connect to QBO
    print("Connecting to QuickBooks...")
    qbo = QBOService(settings.qbo_app, client_name)

    if not qbo.is_authenticated:
        print(f"ERROR: {client_name} not authorized.")
        print(f"Run: python src/main.py --auth {client_name}")
        return 1

    if not qbo.test_connection():
        print(f"ERROR: Failed to connect to QBO for {client_name}")
        print(f"Token may be expired. Run: python src/main.py --auth {client_name}")
        return 1

    print("Connected!\n")

    # Fetch Chart of Accounts for injecting zero-balance rows
    print("Fetching Chart of Accounts...")
    all_accounts = qbo.get_accounts()
    if all_accounts:
        print(f"  {len(all_accounts)} active accounts found\n")
    else:
        print("  WARNING: Could not fetch COA, reports will only show accounts with activity\n")

    # Create export directory
    export_dir = get_export_dir(output_dir)
    print(f"Export directory: {export_dir}\n")

    # Fetch and export each report
    results: List[Tuple[str, str, int, str]] = []

    for report_name in report_names:
        print(f"Fetching: {report_name}...")

        try:
            report_data = qbo.get_report(
                report_name=report_name,
                year=year,
                display=display,
                basis=basis,
                full_year=True,
            )
        except Exception as e:
            print(f"  ERROR: {e}")
            logger.exception(f"Failed to fetch {report_name}")
            results.append((report_name, "FAILED", 0, ""))
            continue

        if not report_data:
            print(f"  ERROR: No data returned for {report_name}")
            results.append((report_name, "FAILED", 0, ""))
            continue

        # Inject zero-balance accounts from COA (P&L and Balance Sheet only)
        if all_accounts and report_name in (
            "Balance Sheet", "P&L", "Profit and Loss",
        ):
            qbo.inject_missing_accounts(report_data, all_accounts)

        # Parse with MAX_ALL to get everything
        rows, headers = qbo.parse_report_to_rows(
            report_data,
            row_max=MAX_ALL,
            col_max=MAX_ALL,
        )

        if not rows and not headers:
            print(f"  WARNING: Empty report for {report_name}")
            results.append((report_name, "EMPTY", 0, ""))
            continue

        # Write to Excel
        filename = generate_filename(client_name, report_name)
        filepath = export_dir / filename

        num_rows = write_to_excel(filepath, headers, rows, report_name)
        print(f"  Saved: {filename} ({num_rows} rows, {len(headers)} columns)")
        results.append((report_name, "OK", num_rows, filename))

    # Print summary
    error_count = sum(1 for _, status, _, _ in results if status == "FAILED")

    print(f"\n{'=' * 60}")
    print("  SUMMARY")
    print(f"{'=' * 60}")
    for rname, status, nrows, fname in results:
        if status == "OK":
            print(f"  OK    {rname:30s} {nrows:>5d} rows -> {fname}")
        elif status == "EMPTY":
            print(f"  EMPTY {rname:30s}")
        else:
            print(f"  FAIL  {rname:30s}")

    total = len(results)
    ok_count = total - error_count
    print(f"\n  {ok_count}/{total} reports exported successfully.")
    if error_count:
        print(f"  {error_count} report(s) failed.")
    print()

    return 0 if error_count == 0 else 1


def main():
    """Entry point."""
    parser = argparse.ArgumentParser(
        description="QBO Test Export - Download QBO reports to Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python src/test_qbo_export.py --client BostonHCP
  python src/test_qbo_export.py --client BostonHCP --report "Balance Sheet" --report "P&L"
  python src/test_qbo_export.py --client BostonHCP --year 2025 --basis Cash
  python src/test_qbo_export.py --list-clients
  python src/test_qbo_export.py --list-reports
        """,
    )

    parser.add_argument("--client", help="Client name to export reports for")
    parser.add_argument(
        "--report", action="append", dest="reports",
        help="Report name (repeatable). Defaults to 4 main reports.",
    )
    parser.add_argument(
        "--all-reports", action="store_true",
        help="Fetch all known report types",
    )
    parser.add_argument(
        "--year", type=int, default=date.today().year,
        help=f"Report year (default: {date.today().year})",
    )
    parser.add_argument(
        "--display", default="Monthly",
        choices=["Monthly", "Weekly", "Quarterly", "Yearly"],
        help="Report display type (default: Monthly)",
    )
    parser.add_argument(
        "--basis", default="Accrual",
        choices=["Cash", "Accrual"],
        help="Accounting basis (default: Accrual)",
    )
    parser.add_argument(
        "--output-dir", default="exports",
        help="Output directory relative to project root (default: exports)",
    )
    parser.add_argument("--list-clients", action="store_true", help="List configured clients")
    parser.add_argument("--list-reports", action="store_true", help="List available report types")

    args = parser.parse_args()

    # Console-only logging for test script
    setup_logger(log_to_file=False)

    # Handle list commands
    if args.list_clients:
        list_clients()
        return

    if args.list_reports:
        list_reports()
        return

    # Require --client for export
    if not args.client:
        parser.error("--client is required (or use --list-clients / --list-reports)")

    # Determine reports to fetch
    if args.all_reports:
        seen_api_names = set()
        report_names = []
        for display_name, api_name in QBO_REPORTS.items():
            if api_name not in seen_api_names:
                seen_api_names.add(api_name)
                report_names.append(display_name)
    elif args.reports:
        report_names = args.reports
    else:
        report_names = DEFAULT_REPORTS

    exit_code = export_reports(
        client_name=args.client,
        report_names=report_names,
        year=args.year,
        display=args.display,
        basis=args.basis,
        output_dir=args.output_dir,
    )
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
